from openpyxl.reader.excel import load_workbook

from common.utils.log import log

module_ram_path = "resource/register_ddic.xlsx"
module_ram_add_path = "resource/register_ddic.xlsx"


def readSheetByName_outputDic(row_num1, row_num2, module_sheet_name):
    workbook = load_workbook(filename=module_ram_path)
    # 创建一个空字典来存储数据
    address_dict = {}
    sheet = workbook[module_sheet_name]
    # max_column获取具有数据的最大列序
    max_column = sheet.max_column
    # ASCII转码
    str1 = chr(max_column + 64)
    key = 1
    # 从第二行开始读取sheet并将其中指定的两列数据保存为字典的key与value
    for row in sheet[f'A2:{str1}' + str(sheet.max_row)]:
        if row_num1 is not None:
            if row_num2 is not None:
                value2 = row[row_num2].value
            else:
                value2 = None
            address_dict[key] = [row[row_num1].value, value2]
        else:
            raise ValueError(f"{row_num1}为None")
        key += 1
    return address_dict


def readSheetByName_outputList(row_num, module_sheet_name):
    workbook = load_workbook(filename=module_ram_path)
    # 创建一个空列表来存储数据
    address_list = []
    sheet = workbook[module_sheet_name]
    # max_column获取具有数据的最大列序
    max_column = sheet.max_column
    str1 = chr(max_column + 64)
    # 从第二行开始读取sheet并将其中指定的两列数据保存为字典的key与value
    for row in sheet[f'A2:{str1}' + str(sheet.max_row)]:
        if row_num is not None:
            address_list.append(row(row_num).value)
        else:
            raise ValueError(f"{row_num}为None")

    return address_list


# 通过输入的两个值获取指定sheet的指定表头在该sheet的表头list中的index
def find_headerNum(table_header, module_sheet_name):
    if table_header:
        workbook = load_workbook(module_ram_path)
        sheet = workbook[module_sheet_name]
        column_names = [cell.value for col in sheet.iter_cols() for cell in col if cell.row == 1]

        if table_header in column_names:
            return column_names.index(table_header)
        else:
            print(f"指定sheet中没有{table_header}列")
            return None
            # raise ValueError(f"指定sheet中没有{table_header}列")
    else:
        print(f"find_tableHeader输入参数为{table_header}")
        return None


# 将list顺序保存到指定sheet的列中
def saveDicToExcel(input_list, module_sheet_name):
    # 加载工作簿
    workbook = load_workbook(module_ram_path)
    sheet = workbook[module_sheet_name]  # 获取活动工作表
    # 获取最后一列的列号
    raw_last_column = sheet.max_column
    header_num = find_headerNum(input_list[0], module_sheet_name)
    if header_num is None:
        log.info(f"在sheet尾部新增一列数据，列名：{input_list[0]}")
        target_column = raw_last_column + 1
    else:
        target_column = header_num + 1
    # Row or column values must be at least 1
    for i in range(1, len(input_list) + 1):
        sheet.cell(row=i, column=target_column, value=input_list[i - 1])
    # 保存更改后的工作簿
    workbook.save(module_ram_add_path)
    return True


# 获取excel的全部sheet名并输出为list
def getExcellAllSheetName():
    workbook = load_workbook(module_ram_path)
    sheet_name_list = workbook.sheetnames
    return sheet_name_list


# 获取包含字符的两个行的行数
def selectInvalidRow(input_sheet):
    signs = ['/']
    workbook = load_workbook(filename=module_ram_path)
    sheet = workbook[input_sheet]
    output_list = []

    invalid_list = [cell.value for cell in sheet['A']]
    for i in range(len(invalid_list)):
        if invalid_list[i] in signs:
            output_list.append(i + 1)
    print(output_list)
    return output_list


def readSheetByName_outputDic_effectRow(row_num1, row_num2, module_sheet_name):
    workbook = load_workbook(filename=module_ram_path)
    sheet = workbook[module_sheet_name]
    invalid_list = selectInvalidRow(module_sheet_name)

    # 执行读取值
    # 创建一个空字典来存储数据
    address_dict = {}

    for row in sheet.iter_rows(min_row = 2, max_col=sheet.max_column,values_only=True):
        address_dict[row[row_num1]] = row[row_num2]

    return address_dict
 

# 将list顺序保存到指定sheet的列中
def saveDicToExcel_effectRow(input_list, module_sheet_name):
    str_sign = '/'
    # 加载工作簿
    workbook = load_workbook(module_ram_path)
    sheet = workbook[module_sheet_name]  # 获取活动工作表
    # 获取最后一列的列号
    raw_last_column = sheet.max_column
    header_num = find_headerNum(input_list[0], module_sheet_name)
    # 判断是否存在需要保存的列
    if header_num is None:
        log.info(f"在sheet尾部新增一列数据，列名：{input_list[0]}")
        target_column = raw_last_column + 1
    else:
        target_column = header_num + 1

    target_raw = 1
    # 行列从1开始，以A列为判断依据，如果有标识符/则保存时跳过该行
    for i in range(1, sheet.max_row + 1):
        for row in sheet.iter_rows(min_row=i, max_row=i, min_col=1, max_col=sheet.max_column):
            if str_sign not in row[0].value:
                sheet.cell(row=i, column=target_column, value=input_list[target_raw - 1])
                target_raw += 1

    # 保存更改后的工作簿
    workbook.save(module_ram_add_path)
    return True

def getrowvalue(module_sheet_name:str, rownumber):
    worksheet = load_workbook(module_ram_path)
    sheet = worksheet[module_sheet_name]  # 获取活动工作表

    totalrows = sheet.max_row

    row_data = []
    for cell in sheet[rownumber]:
        row_data.append(cell.value)

    return  row_data

def getmaxval(module_sheet_name:str):
    worksheet = load_workbook(module_ram_path)
    sheet = worksheet[module_sheet_name]  # 获取活动工作表

    return sheet.max_row